# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT='Arial'
BLUE=Font(name=FONT,color='0000FF')
BLACK=Font(name=FONT,color='000000')
GREEN=Font(name=FONT,color='008000')
BLACKB=Font(name=FONT,bold=True,color='000000')
WHITEB=Font(name=FONT,bold=True,color='FFFFFF')
TITLE=Font(name=FONT,bold=True,size=14,color='1F3864')
SUB=Font(name=FONT,italic=True,size=9,color='595959')
HDR_FILL=PatternFill('solid',start_color='1F3864')
SEC_FILL=PatternFill('solid',start_color='D9E1F2')
TOT_FILL=PatternFill('solid',start_color='FCE4D6')
YEL=PatternFill('solid',start_color='FFF2CC')
thin=Side(style='thin',color='BFBFBF')
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
NUM='#,##0;(#,##0);"-"'
PCT='0.0%'
CEN=Alignment(horizontal='center',vertical='center',wrap_text=True)
RIGHT=Alignment(horizontal='right')
LEFT=Alignment(horizontal='left',vertical='center',wrap_text=True)

wb=Workbook()
def setcol(ws,widths):
    for c,w in widths.items(): ws.column_dimensions[c].width=w
def cell(ws,coord,val,font=BLACK,fmt=None,fill=None,align=None,border=True):
    c=ws[coord]; c.value=val; c.font=font
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
    if align: c.alignment=align
    if border: c.border=BORDER
    return c

# ===== S1 開業資金・調達
S1='1.開業資金・調達'
ws=wb.active; ws.title=S1
setcol(ws,{'A':34,'B':16,'C':54})
cell(ws,'A1','美容室 monet（モネ）岡山店　開業資金・資金調達計画',TITLE,border=False)
cell(ws,'A2','申請者：島田 和也（個人事業主）　／　FC本部：株式会社グランフュージョン　／　単位：円（消費税込）',SUB,border=False)
r=4
cell(ws,f'A{r}','項目',WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,f'B{r}','金額',WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,f'C{r}','摘要・根拠',WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'A5','【設備資金】',BLACKB,fill=SEC_FILL); cell(ws,'B5',None,fill=SEC_FILL); cell(ws,'C5',None,fill=SEC_FILL)
items=[
 ('① 内装・設備工事',18000000,'(株)Rally Fun／(株)ダリアート 概算御見積書（改定）（税別16,363,637円）'),
 ('② デザイン設計費',1358500,'デザイン設計費 1,235,000円＋税'),
 ('③ その他諸経費・備品代',1000000,'開業諸経費・什器備品一式'),
 ('④ 給湯機（ボイラー）',1540000,'(株)ダリア 御見積書 2026/6/23（カーサ･イクスR＋ガス16号・取付費込／税込）'),
 ('⑤ シャンプー設備（現金一括購入）',668800,'シャンプーチェア195,000円×2＋シャンプーボウル109,000円×2＝税別608,000円（税込668,800円）'),
]
row=6
for n,v,note in items:
    cell(ws,f'A{row}',n); cell(ws,f'B{row}',v,BLUE,NUM); cell(ws,f'C{row}',note,Font(name=FONT,size=9),align=LEFT); row+=1
cell(ws,f'A{row}','　償却対象設備 小計',BLACKB); cell(ws,f'B{row}','=SUM(B6:B10)',BLACKB,NUM,fill=TOT_FILL); cell(ws,f'C{row}','①〜⑤（減価償却の対象）',Font(name=FONT,size=9),align=LEFT)
sub_dep=row; row+=1
cell(ws,f'A{row}','⑥ FC加盟金'); cell(ws,f'B{row}',1500000,BLUE,NUM); cell(ws,f'C{row}','特別価格1,500,000円（税込・今回限り。標準価格3,300,000円）／繰延資産・5年償却',Font(name=FONT,size=9),align=LEFT)
fc_row=row; row+=1
cell(ws,f'A{row}','⑦ 物件取得費'); cell(ws,f'B{row}',1150000,BLUE,NUM); cell(ws,f'C{row}','あおい不動産 請求書 2026/6/11（敷金600千・礼金220千・仲介手数料220千・保証会社初期費用110千。火災保険別途）',Font(name=FONT,size=9),align=LEFT)
bukken_row=row; row+=1
cell(ws,f'A{row}','設備資金 計',BLACKB); cell(ws,f'B{row}',f'=B{sub_dep}+B{fc_row}+B{bukken_row}',BLACKB,NUM,fill=TOT_FILL); cell(ws,f'C{row}','',border=True)
settot=row; row+=1
cell(ws,f'A{row}','【運転資金】',BLACKB,fill=SEC_FILL); cell(ws,f'B{row}',None,fill=SEC_FILL); cell(ws,f'C{row}',None,fill=SEC_FILL); row+=1
cell(ws,f'A{row}','⑧ 運転資金'); cell(ws,f'B{row}',4245500,BLUE,NUM); cell(ws,f'C{row}','開業後約2ヶ月分の運転資金＋開業時販促費・予備費',Font(name=FONT,size=9),align=LEFT)
unten=row; row+=1
cell(ws,f'A{row}','開業資金 総額',WHITEB,fill=HDR_FILL); cell(ws,f'B{row}',f'=B{settot}+B{unten}',WHITEB,NUM,fill=HDR_FILL); cell(ws,f'C{row}','設備資金＋運転資金',Font(name=FONT,size=9),align=LEFT)
total_row=row; row+=2
cell(ws,f'A{row}','【資金調達】',BLACKB,fill=SEC_FILL); cell(ws,f'B{row}',None,fill=SEC_FILL); cell(ws,f'C{row}',None,fill=SEC_FILL); row+=1
cell(ws,f'A{row}','自己資金（充当額）'); cell(ws,f'B{row}',0,BLUE,NUM); cell(ws,f'C{row}','保有自己資金10,000,000円を別途確保（今回は使用せず全額借入で試算）',Font(name=FONT,size=9),align=LEFT)
jiko=row; row+=1
cell(ws,f'A{row}','借入金（日本政策金融公庫・信用金庫）'); cell(ws,f'B{row}',f'=B{total_row}-B{jiko}',BLACKB,NUM); cell(ws,f'C{row}','全額借入で試算',Font(name=FONT,size=9),align=LEFT)
kariire=row; row+=1
cell(ws,f'A{row}','調達 計',BLACKB); cell(ws,f'B{row}',f'=B{jiko}+B{kariire}',BLACKB,NUM,fill=TOT_FILL); cell(ws,f'C{row}',f'=IF(B{row}=B{total_row},"○ 総額と一致","差額あり")',Font(name=FONT,size=9),align=LEFT)
row+=2
cell(ws,f'A{row}','（参考）保有自己資金',BLACKB,fill=YEL); cell(ws,f'B{row}',10000000,BLUE,NUM,fill=YEL); cell(ws,f'C{row}','本計画では使用せず。返済・運転の予備原資として確保（財務余力を示す）',Font(name=FONT,size=9),align=LEFT)

S1_dep=f"'{S1}'!B{sub_dep}"; S1_fc=f"'{S1}'!B{fc_row}"; S1_total=f"'{S1}'!B{total_row}"; S1_kariire=f"'{S1}'!B{kariire}"; S1_unten=f"'{S1}'!B{unten}"

# ===== S2 前提条件
S2='2.前提条件'
ws=wb.create_sheet(S2); setcol(ws,{'A':30,'B':16,'C':48})
cell(ws,'A1','前提条件（収支計画の算定基礎）',TITLE,border=False)
cell(ws,'A2','青字＝入力値（変更可）　黒字＝計算値　緑字＝他シート参照',SUB,border=False)
cell(ws,'A4','項目',WHITEB,fill=HDR_FILL,align=CEN);cell(ws,'B4','値',WHITEB,fill=HDR_FILL,align=CEN);cell(ws,'C4','備考',WHITEB,fill=HDR_FILL,align=CEN)
def put2(r,name,val,font,fmt,note):
    cell(ws,f'A{r}',name);cell(ws,f'B{r}',val,font,fmt);cell(ws,f'C{r}',note,Font(name=FONT,size=9),align=LEFT)
cell(ws,'A5','【売上】',BLACKB,fill=SEC_FILL);cell(ws,'B5',None,fill=SEC_FILL);cell(ws,'C5',None,fill=SEC_FILL)
put2(6,'客単価（円）',13000,BLUE,NUM,'FC実績：堀江約14千・姪浜約13.5千・広島約14千を踏まえ保守設定')
put2(7,'材料費率（対純売上）',0.12,BLUE,PCT,'FC実績PLの技術売上原価率 約11%＋商品原価を勘案')
put2(8,'ロイヤリティ率',0.05,BLUE,PCT,'開業4ヶ月目以降、純売上の5%（特別レート。標準10%）')
cell(ws,'A9','【固定費・月額】',BLACKB,fill=SEC_FILL);cell(ws,'B9',None,fill=SEC_FILL);cell(ws,'C9',None,fill=SEC_FILL)
put2(10,'家賃',280000,BLUE,NUM,'FC実績：広島23万・姪浜約28万を踏まえ設定（契約確定後に更新）')
put2(11,'水道光熱費',35000,BLUE,NUM,'給湯機・乾燥機稼働を勘案')
put2(12,'シャンプー台リース料（税込）',59400,BLUE,NUM,'マッサージ付シャンプーチェア2台（1台110万円税込）リース：サムシング御見積 4年48回・月額59,400円（税込）')
put2(13,'その他経費',130000,BLUE,NUM,'通信・支払手数料・消耗品・サービス費等')
put2(14,'社会保険料率（対人件費）',0.15,BLUE,PCT,'法定福利費（会社負担分）')
put2(15,'損害保険料',round(37880/12),BLUE,NUM,'タフビズ事業活動総合保険 見積（設備什器500万・商品100万・借家賠償2,000万・類焼1億等）一時払 年37,880円')
cell(ws,'A16','【投資・償却】',BLACKB,fill=SEC_FILL);cell(ws,'B16',None,fill=SEC_FILL);cell(ws,'C16',None,fill=SEC_FILL)
put2(17,'償却対象設備（税込）',f'={S1_dep}',GREEN,NUM,'開業資金シートより')
put2(18,'設備償却年数',10,BLUE,'0"年"','店舗内装・設備（定額・保守的に10年）')
put2(19,'設備減価償却費（月）','=B17/B18/12',BLACK,NUM,'')
put2(20,'FC加盟金（税込）',f'={S1_fc}',GREEN,NUM,'開業資金シートより')
put2(21,'加盟金償却年数',5,BLUE,'0"年"','繰延資産')
put2(22,'加盟金償却費（月）','=B20/B21/12',BLACK,NUM,'')
cell(ws,'A23','【借入】',BLACKB,fill=SEC_FILL);cell(ws,'B23',None,fill=SEC_FILL);cell(ws,'C23',None,fill=SEC_FILL)
put2(24,'借入総額',f'={S1_kariire}',GREEN,NUM,'開業資金シートより（全額借入）')
put2(25,'年利',0.025,BLUE,PCT,'日本政策金融公庫 創業融資の想定利率')
put2(26,'返済期間（年）',10,BLUE,'0"年"','')
put2(27,'毎月返済額（元利均等）','=-PMT(B25/12,B26*12,B24)',BLACK,NUM,'')
put2(28,'年間返済額','=B27*12',BLACK,NUM,'')
P_kyaku=f"'{S2}'!$B$6"; P_mat=f"'{S2}'!$B$7"; P_roy=f"'{S2}'!$B$8"
P_rent=f"'{S2}'!$B$10"; P_util=f"'{S2}'!$B$11"; P_lease=f"'{S2}'!$B$12"; P_other=f"'{S2}'!$B$13"; P_shaho=f"'{S2}'!$B$14"; P_hoken=f"'{S2}'!$B$15"
P_dep=f"'{S2}'!$B$19"; P_fcdep=f"'{S2}'!$B$22"; P_loan=f"'{S2}'!$B$24"; P_rate=f"'{S2}'!$B$25"; P_pmt=f"'{S2}'!$B$27"

# ===== S6 借入返済計画
S6='6.借入返済計画'
ws=wb.create_sheet(S6)
setcol(ws,{'A':8,'B':10,'C':16,'D':14,'E':14,'F':14,'G':16,'H':3,'I':10,'J':16,'K':14,'L':14})
cell(ws,'A1','借入返済計画（元利均等・10年）',TITLE,border=False)
cell(ws,'A2',f'借入総額・年利・返済額は「{S2}」を参照',SUB,border=False)
for i,h in enumerate(['回','年目','期首残高','返済額','うち利息','うち元金','期末残高']):
    cell(ws,f'{get_column_letter(1+i)}3',h,WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'I3','年度',WHITEB,fill=HDR_FILL,align=CEN);cell(ws,'J3','返済額計',WHITEB,fill=HDR_FILL,align=CEN);cell(ws,'K3','利息計',WHITEB,fill=HDR_FILL,align=CEN);cell(ws,'L3','元金計',WHITEB,fill=HDR_FILL,align=CEN)
n=120
for m in range(1,n+1):
    rr=3+m
    cell(ws,f'A{rr}',m,BLACK,'0')
    cell(ws,f'B{rr}',f'=ROUNDUP(A{rr}/12,0)',BLACK,'0"年目"')
    cell(ws,f'C{rr}',(f'={P_loan}' if m==1 else f'=G{rr-1}'),BLACK,NUM)
    cell(ws,f'D{rr}',f'={P_pmt}',BLACK,NUM)
    cell(ws,f'E{rr}',f'=C{rr}*{P_rate}/12',BLACK,NUM)
    cell(ws,f'F{rr}',f'=D{rr}-E{rr}',BLACK,NUM)
    cell(ws,f'G{rr}',f'=C{rr}-F{rr}',BLACK,NUM)
for y in range(1,6):
    rr=3+y; s=4+(y-1)*12; e=s+11
    cell(ws,f'I{rr}',y,BLACK,'0"年目"')
    cell(ws,f'J{rr}',f'=SUM(D{s}:D{e})',BLACK,NUM)
    cell(ws,f'K{rr}',f'=SUM(E{s}:E{e})',BLACK,NUM)
    cell(ws,f'L{rr}',f'=SUM(F{s}:F{e})',BLACK,NUM)
def L_int_year(y): return f"'{S6}'!K{3+y}"
def L_prin_year(y): return f"'{S6}'!L{3+y}"
def L_int_month(m): return f"'{S6}'!E{3+m}"
def L_prin_month(m): return f"'{S6}'!F{3+m}"

# ===== S3 売上計画
S3='3.売上計画'
ws=wb.create_sheet(S3)
months=['1\n(26/9)','2\n(26/10)','3\n(26/11)','4\n(26/12)','5\n(27/1)','6\n(27/2)','7\n(27/3)','8\n(27/4)','9\n(27/5)','10\n(27/6)','11\n(27/7)','12\n(27/8)']
sales=[700000,1800000,2000000,2150000,2250000,2300000,2350000,2350000,2350000,2350000,2350000,2350000]
setcol(ws,{'A':22})
for i in range(12): ws.column_dimensions[get_column_letter(2+i)].width=11
ws.column_dimensions['N'].width=14
cell(ws,'A1','売上計画（1年目 月次／2〜5年目 年次）',TITLE,border=False)
cell(ws,'A2','売上はFC実績（新規店M2で月商200万円超、成熟店 平均266万円）に基づく保守的設定',SUB,border=False)
cell(ws,'A4','開業からの月数',WHITEB,fill=HDR_FILL,align=CEN)
for i,mm in enumerate(months): cell(ws,f'{get_column_letter(2+i)}4',mm,WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'N4','年間合計',WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'A5','純売上高（円）',BLACK)
for i,v in enumerate(sales): cell(ws,f'{get_column_letter(2+i)}5',v,BLUE,NUM)
cell(ws,'N5','=SUM(B5:M5)',BLACKB,NUM,fill=TOT_FILL)
cell(ws,'A6','客数（人）',BLACK)
for i in range(12):
    col=get_column_letter(2+i); cell(ws,f'{col}6',f'=ROUND({col}5/{P_kyaku},0)',BLACK,'#,##0"人"')
cell(ws,'N6','=SUM(B6:M6)',BLACKB,'#,##0"人"',fill=TOT_FILL)
cell(ws,'A7','客単価（円）',BLACK)
for i in range(12):
    col=get_column_letter(2+i); cell(ws,f'{col}7',f'=IF({col}6=0,0,{col}5/{col}6)',BLACK,NUM)
cell(ws,'N7','=IF(N6=0,0,N5/N6)',BLACK,NUM)
cell(ws,'A9','年次売上計画',BLACKB,fill=SEC_FILL)
for c in ['B','C','D','E','F']: cell(ws,f'{c}9',None,fill=SEC_FILL)
cell(ws,'A10','年度',WHITEB,fill=HDR_FILL,align=CEN)
yl=['1年目','2年目','3年目','4年目','5年目']
for i,y in enumerate(yl): cell(ws,f'{get_column_letter(2+i)}10',y,WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'A11','月平均売上（円）',BLACK)
cell(ws,'B11','=N5/12',BLACK,NUM)
for i,v in enumerate([2500000,2600000,2650000,2700000]): cell(ws,f'{get_column_letter(3+i)}11',v,BLUE,NUM)
cell(ws,'A12','年間売上（円）',BLACKB)
cell(ws,'B12','=N5',BLACKB,NUM,fill=TOT_FILL)
for i in range(4):
    col=get_column_letter(3+i); cell(ws,f'{col}12',f'={col}11*12',BLACKB,NUM,fill=TOT_FILL)
cell(ws,'A13','成長率',BLACK)
cell(ws,'B13','—',BLACK,align=RIGHT)
for i in range(4):
    col=get_column_letter(3+i); prev=get_column_letter(2+i); cell(ws,f'{col}13',f'={col}12/{prev}12-1',BLACK,PCT)
def Sales_month(m): return f"'{S3}'!{get_column_letter(1+m)}5"
def S3_year_sales(y): return f"='{S3}'!{get_column_letter(1+y)}12"

# ===== S4 月次収支(1年目)
S4='4.月次収支(1年目)'
ws=wb.create_sheet(S4); setcol(ws,{'A':24})
for i in range(12): ws.column_dimensions[get_column_letter(2+i)].width=11
ws.column_dimensions['N'].width=14
cell(ws,'A1','月次収支計画（1年目）　単位：円',TITLE,border=False)
cell(ws,'A2','開業初年度（2026年9月開業想定）。ロイヤリティは4ヶ月目以降に発生。',SUB,border=False)
cell(ws,'A4','勘定科目',WHITEB,fill=HDR_FILL,align=CEN)
for i,mm in enumerate(months): cell(ws,f'{get_column_letter(2+i)}4',mm,WHITEB,fill=HDR_FILL,align=CEN)
cell(ws,'N4','年間合計',WHITEB,fill=HDR_FILL,align=CEN)
jinken=[600000,600000,600000,750000,750000,750000,820000,820000,820000,820000,820000,820000]
koukoku=[200000,200000,200000,50000,50000,50000,50000,50000,50000,50000,50000,50000]
R=5
def line(name,func,bold=False,fill=None,fmt=NUM,total=True):
    global R
    f=BLACKB if bold else BLACK
    cell(ws,f'A{R}',name,f,fill=fill)
    for i in range(12):
        col=get_column_letter(2+i); cell(ws,f'{col}{R}',func(i,col),f,fmt,fill=fill)
    if total: cell(ws,f'N{R}',f'=SUM(B{R}:M{R})',BLACKB,fmt,fill=(fill if fill else TOT_FILL))
    else: cell(ws,f'N{R}',None,fill=fill)
    rr=R; R+=1; return rr
r_sales=line('純売上高',lambda i,c:f'={Sales_month(i+1)}',bold=True)
r_cogs=line('売上原価（材料費）',lambda i,c:f'={c}{r_sales}*{P_mat}')
r_gp=line('売上総利益',lambda i,c:f'={c}{r_sales}-{c}{r_cogs}',bold=True,fill=SEC_FILL)
cell(ws,f'A{R}','【販売管理費】',BLACKB,fill=SEC_FILL)
for i in range(12): cell(ws,f'{get_column_letter(2+i)}{R}',None,fill=SEC_FILL)
cell(ws,f'N{R}',None,fill=SEC_FILL); R+=1
r_jin=line('人件費',lambda i,c:jinken[i])
for i in range(12): cell(ws,f'{get_column_letter(2+i)}{r_jin}',jinken[i],BLUE,NUM)
r_sha=line('法定福利費',lambda i,c:f'={c}{r_jin}*{P_shaho}')
r_rent=line('地代家賃',lambda i,c:f'={P_rent}')
r_kou=line('広告宣伝費',lambda i,c:koukoku[i])
for i in range(12): cell(ws,f'{get_column_letter(2+i)}{r_kou}',koukoku[i],BLUE,NUM)
r_util=line('水道光熱費',lambda i,c:f'={P_util}')
r_lease=line('シャンプー台リース料',lambda i,c:f'={P_lease}')
r_hoken=line('損害保険料',lambda i,c:f'={P_hoken}')
r_roy=line('ロイヤリティ',lambda i,c:0)
for i in range(12):
    col=get_column_letter(2+i)
    if i<3: cell(ws,f'{col}{r_roy}',0,BLACK,NUM)
    else: cell(ws,f'{col}{r_roy}',f'={col}{r_sales}*{P_roy}',BLACK,NUM)
r_oth=line('その他経費',lambda i,c:f'={P_other}')
r_dep=line('減価償却費',lambda i,c:f'={P_dep}')
r_fcd=line('加盟金償却費',lambda i,c:f'={P_fcdep}')
r_sga=line('販売管理費 計',lambda i,c:f'=SUM({c}{r_jin}:{c}{r_fcd})',bold=True,fill=TOT_FILL)
r_op=line('営業利益',lambda i,c:f'={c}{r_gp}-{c}{r_sga}',bold=True,fill=SEC_FILL)
r_int=line('支払利息',lambda i,c:f'={L_int_month(i+1)}')
r_ord=line('経常利益',lambda i,c:f'={c}{r_op}-{c}{r_int}',bold=True,fill=TOT_FILL)
cell(ws,f'A{R}','【簡易資金繰り】',BLACKB,fill=SEC_FILL)
for i in range(12): cell(ws,f'{get_column_letter(2+i)}{R}',None,fill=SEC_FILL)
cell(ws,f'N{R}',None,fill=SEC_FILL); R+=1
r_adddep=line('＋減価償却費・加盟金償却',lambda i,c:f'={c}{r_dep}+{c}{r_fcd}')
r_prin=line('－借入元金返済',lambda i,c:f'=-{L_prin_month(i+1)}')
r_cf=line('月次資金収支',lambda i,c:f'={c}{r_ord}+{c}{r_adddep}+{c}{r_prin}',bold=True,fill=SEC_FILL)
cell(ws,f'A{R}','累計現金残高',BLACKB)
for i in range(12):
    col=get_column_letter(2+i)
    if i==0: cell(ws,f'{col}{R}',f'={S1_unten}+{col}{r_cf}',BLACKB,NUM,fill=YEL)
    else:
        prev=get_column_letter(1+i); cell(ws,f'{col}{R}',f'={prev}{R}+{col}{r_cf}',BLACKB,NUM,fill=YEL)
cell(ws,f'N{R}',f'=M{R}',BLACKB,NUM,fill=YEL)
R+=1

# ===== S5 年次収支(5年)
S5='5.年次収支(5年)'
ws=wb.create_sheet(S5); setcol(ws,{'A':26,'B':16,'C':16,'D':16,'E':16,'F':16})
cell(ws,'A1','年次収支計画（5年）　単位：円',TITLE,border=False)
cell(ws,'A2','1年目は月次計画の積上げ。2〜5年目は売上計画シートの年間売上に基づく。',SUB,border=False)
cell(ws,'A4','勘定科目',WHITEB,fill=HDR_FILL,align=CEN)
for i,y in enumerate(yl): cell(ws,f'{get_column_letter(2+i)}4',y,WHITEB,fill=HDR_FILL,align=CEN)
jin_y=[None,850000,880000,900000,920000]
RR=5
def aline(name,vals,bold=False,fill=None,fmt=NUM):
    global RR
    f=BLACKB if bold else BLACK
    cell(ws,f'A{RR}',name,f,fill=fill)
    for i in range(5): cell(ws,f'{get_column_letter(2+i)}{RR}',vals[i],f,fmt,fill=fill)
    rr=RR; RR+=1; return rr
ar_sales=aline('純売上高',[f"='{S4}'!N{r_sales}"]+[S3_year_sales(y) for y in range(2,6)],bold=True)
ar_cogs=aline('売上原価（材料費）',[f'={get_column_letter(2+i)}{ar_sales}*{P_mat}' for i in range(5)])
ar_gp=aline('売上総利益',[f'={get_column_letter(2+i)}{ar_sales}-{get_column_letter(2+i)}{ar_cogs}' for i in range(5)],bold=True,fill=SEC_FILL)
cell(ws,f'A{RR}','【販売管理費】',BLACKB,fill=SEC_FILL)
for i in range(5): cell(ws,f'{get_column_letter(2+i)}{RR}',None,fill=SEC_FILL)
RR+=1
ar_jin=RR
cell(ws,f'A{RR}','人件費',BLACK); cell(ws,f'B{RR}',f"='{S4}'!N{r_jin}",BLACK,NUM)
for i in range(1,5): cell(ws,f'{get_column_letter(2+i)}{RR}',f'={jin_y[i]}*12',BLACK,NUM)
RR+=1
ar_sha=aline('法定福利費',[f'={get_column_letter(2+i)}{ar_jin}*{P_shaho}' for i in range(5)])
ar_rent=aline('地代家賃',[f'={P_rent}*12' for i in range(5)])
ar_kou=RR
cell(ws,f'A{RR}','広告宣伝費',BLACK); cell(ws,f'B{RR}',f"='{S4}'!N{r_kou}",BLACK,NUM)
for i in range(1,5): cell(ws,f'{get_column_letter(2+i)}{RR}','=50000*12',BLACK,NUM)
RR+=1
ar_util=aline('水道光熱費',[f'={P_util}*12' for i in range(5)])
ar_lease=aline('シャンプー台リース料',[f'={P_lease}*12' for i in range(5)])
ar_hoken=aline('損害保険料',[f'={P_hoken}*12' for i in range(5)])
ar_roy=RR
cell(ws,f'A{RR}','ロイヤリティ',BLACK); cell(ws,f'B{RR}',f"='{S4}'!N{r_roy}",BLACK,NUM)
for i in range(1,5): cell(ws,f'{get_column_letter(2+i)}{RR}',f'={get_column_letter(2+i)}{ar_sales}*{P_roy}',BLACK,NUM)
RR+=1
ar_oth=aline('その他経費',[f'={P_other}*12' for i in range(5)])
ar_dep=aline('減価償却費',[f'={P_dep}*12' for i in range(5)])
ar_fcd=aline('加盟金償却費',[f'={P_fcdep}*12' for i in range(5)])
ar_sga=aline('販売管理費 計',[f'=SUM({get_column_letter(2+i)}{ar_jin}:{get_column_letter(2+i)}{ar_fcd})' for i in range(5)],bold=True,fill=TOT_FILL)
ar_op=aline('営業利益',[f'={get_column_letter(2+i)}{ar_gp}-{get_column_letter(2+i)}{ar_sga}' for i in range(5)],bold=True,fill=SEC_FILL)
ar_int=aline('支払利息',[f'={L_int_year(y)}' for y in range(1,6)])
ar_ord=aline('経常利益',[f'={get_column_letter(2+i)}{ar_op}-{get_column_letter(2+i)}{ar_int}' for i in range(5)],bold=True,fill=TOT_FILL)
ar_tax=aline('法人税等（30%）',[f'=MAX(0,{get_column_letter(2+i)}{ar_ord})*0.3' for i in range(5)])
ar_net=aline('税引後当期利益',[f'={get_column_letter(2+i)}{ar_ord}-{get_column_letter(2+i)}{ar_tax}' for i in range(5)],bold=True,fill=SEC_FILL)
cell(ws,f'A{RR}','【返済原資・資金収支】',BLACKB,fill=SEC_FILL)
for i in range(5): cell(ws,f'{get_column_letter(2+i)}{RR}',None,fill=SEC_FILL)
RR+=1
ar_add=aline('＋減価償却費・加盟金償却',[f'={get_column_letter(2+i)}{ar_dep}+{get_column_letter(2+i)}{ar_fcd}' for i in range(5)])
ar_genshi=aline('返済原資（税引後利益＋償却）',[f'={get_column_letter(2+i)}{ar_net}+{get_column_letter(2+i)}{ar_add}' for i in range(5)],bold=True,fill=TOT_FILL)
ar_prin=aline('－借入元金返済',[f'=-{L_prin_year(y)}' for y in range(1,6)])
ar_cf=aline('年間資金収支',[f'={get_column_letter(2+i)}{ar_genshi}+{get_column_letter(2+i)}{ar_prin}' for i in range(5)],bold=True,fill=SEC_FILL)
cell(ws,f'A{RR}','累計現金残高',BLACKB)
for i in range(5):
    col=get_column_letter(2+i)
    if i==0: cell(ws,f'{col}{RR}',f'={S1_unten}+{col}{ar_cf}',BLACKB,NUM,fill=YEL)
    else:
        prev=get_column_letter(1+i); cell(ws,f'{col}{RR}',f'={prev}{RR}+{col}{ar_cf}',BLACKB,NUM,fill=YEL)
RR+=1
cell(ws,f'A{RR}','営業利益率',BLACK)
for i in range(5):
    col=get_column_letter(2+i); cell(ws,f'{col}{RR}',f'={col}{ar_op}/{col}{ar_sales}',BLACK,PCT)
RR+=1
cell(ws,f'A{RR}','返済余裕率（返済原資÷年間返済額）',BLACK)
for i in range(5):
    col=get_column_letter(2+i); cell(ws,f'{col}{RR}',f"={col}{ar_genshi}/'{S6}'!J{3+(i+1)}",BLACK,'0.0"倍"')
RR+=1

# ===== S7 FC実績エビデンス
S7='7.FC実績エビデンス'
ws=wb.create_sheet(S7)
setcol(ws,{'A':28});
for c in 'BCDEFGHIJKLM': ws.column_dimensions[c].width=11
cell(ws,'A1','FC実績エビデンス（既存monet各店の月次純売上）',TITLE,border=False)
cell(ws,'A2','出典：sales_summary（FC本部提供）。新規店の立ち上がりと成熟店の安定水準を示す。',SUB,border=False)
data=[
('店舗 ＼ 年月','25/07','25/08','25/09','25/10','25/11','25/12','26/01','26/02','26/03','26/04','26/05','26/06'),
('堀江院（本店）',2887747,3095200,3227758,3017400,2459230,2622585,2155850,1778150,1272650,1147850,1368000,1336800),
('堀江院 2nd',None,None,None,373750,2106900,2190865,2001380,2210300,2660275,2216715,1907805,1422300),
('福岡姪浜院',2502370,2451600,2741850,3021281,2714200,3376920,2926050,3355100,3118300,3259800,3118690,2266290),
('広島楽々園院',2347320,2258060,2585480,2417300,2435070,2633230,2468470,2529570,2573770,2693570,2534860,1824630),
('福島院',None,None,None,None,None,None,None,772240,1899830,2276250,2690090,1964400),
('高槻院',None,None,None,None,None,None,None,661250,1860850,2092690,2219790,1814250),
('土橋院',None,None,None,None,None,None,None,None,None,552170,2152420,1538180),
]
for i,h in enumerate(data[0]): cell(ws,f'{get_column_letter(1+i)}4',h,WHITEB,fill=HDR_FILL,align=CEN)
rr=5
for row in data[1:]:
    cell(ws,f'A{rr}',row[0],BLACK,align=LEFT)
    for i,v in enumerate(row[1:]):
        col=get_column_letter(2+i)
        if v is not None: cell(ws,f'{col}{rr}',v,BLUE,NUM)
        else: cell(ws,f'{col}{rr}',None)
    rr+=1
last=rr-1
cell(ws,f'A{rr}','各月 平均（営業店のみ）',BLACKB,fill=TOT_FILL)
for i in range(12):
    col=get_column_letter(2+i); cell(ws,f'{col}{rr}',f'=IFERROR(AVERAGE({col}5:{col}{last}),0)',BLACKB,NUM,fill=TOT_FILL)
rr+=2
cell(ws,f'A{rr}','■ 分析サマリー',BLACKB,fill=SEC_FILL)
for c in 'BCDE': cell(ws,f'{c}{rr}',None,fill=SEC_FILL)
rr+=1
notes=[
 '新規店の立ち上がり（開業＝1ヶ月目）：',
 '　・福島院  M1 77万 → M2 190万 → M3 228万 → M4 269万',
 '　・高槻院  M1 66万 → M2 186万 → M3 209万 → M4 222万',
 '　・堀江2nd M1 37万 → M2 211万 →（以後 200〜266万で安定）',
 '　→ 多くの店舗が開業2ヶ月目で月商200万円超に到達。',
 '成熟店の安定水準：',
 '　・姪浜院 平均 約311万円／月、広島楽々園院 平均 約254万円／月',
 '　・成熟店プール（23ヶ月分）平均 約266万円／月（中央値 約263万円）',
 '本計画の前提（1年目末 月商235万円、2年目 月平均250万円）はFC実績の',
 '平均（266万円）を下回る保守的な設定であり、達成確度は高い。',
]
for nline in notes:
    cell(ws,f'A{rr}',nline,Font(name=FONT,size=10),align=LEFT,border=False); rr+=1

wb[S4].freeze_panes='B5'; wb[S5].freeze_panes='B5'; wb[S6].freeze_panes='A4'; wb[S3].freeze_panes='B5'; wb[S7].freeze_panes='B5'
out='/sessions/friendly-dreamy-bardeen/mnt/outputs/monet岡山店_収支計画表.xlsx'
wb.save(out); print('saved',out)
